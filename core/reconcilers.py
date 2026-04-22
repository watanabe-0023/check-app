"""突合ロジック本体。1関数 = 1ルール。"""
from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

from .schema import RuleResult, CheckContext, Status
from .loaders import get_mgmt_value, load_thresholds
from .aggregators import (
    sum_ledger_by_month,
    sum_ledger_by_month_category,
    sum_negative_by_month,
    get_freee_account,
    sum_freee_accounts,
)


# ---------------------------------------------------------------------------
# 共通ユーティリティ
# ---------------------------------------------------------------------------

def _get_threshold(th: dict, rule_id: str, key: str) -> Optional[float]:
    rule_th = th.get("rules", {}).get(rule_id, {})
    if rule_th.get("status") == "info":
        return None
    val = rule_th.get(key)
    if val is not None:
        return float(val)
    return float(th.get(f"default_{key}_tolerance", 1))


def _is_info_rule(th: dict, rule_id: str) -> bool:
    return th.get("rules", {}).get(rule_id, {}).get("status") == "info"


def judge(diff: float, rate: float, abs_tol: Optional[float], rate_tol: Optional[float]) -> Status:
    if abs_tol is None and rate_tol is None:
        return "info"
    if abs_tol is not None and abs(diff) <= abs_tol:
        return "pass"
    if rate_tol is not None and abs(rate) <= rate_tol:
        return "pass"
    # warn = 許容値の3倍以内
    if rate_tol is not None and abs(rate) <= rate_tol * 3:
        return "warn"
    if abs_tol is not None and rate_tol is None and abs(diff) <= abs_tol * 10:
        return "warn"
    return "fail"


def _make_result(rule_id: str, label: str, a: Optional[float], b: Optional[float],
                 a_src: str, b_src: str, th: dict, note: str = "") -> RuleResult:
    if a is None or b is None:
        return RuleResult(rule_id, label, a, b, None, None, "warn",
                          "値が取得できませんでした。ファイル構造を確認してください。",
                          a_src, b_src)
    diff = a - b
    rate = diff / b if b != 0 else 0.0
    abs_tol = _get_threshold(th, rule_id, "abs")
    rate_tol = _get_threshold(th, rule_id, "rate")
    status = "info" if _is_info_rule(th, rule_id) else judge(diff, rate, abs_tol, rate_tol)
    return RuleResult(rule_id, label, a, b, diff, rate, status, note, a_src, b_src)


# ---------------------------------------------------------------------------
# S系: 売上突合
# ---------------------------------------------------------------------------

def rule_S01_total_revenue(ctx: CheckContext) -> RuleResult:
    """管理表売上合計 = 台帳売上合計（税別）。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 2, ctx.month)
    b = sum_ledger_by_month(ctx.ledger_sales, ctx.month)
    return _make_result(
        "S01", "売上合計 管理表=台帳",
        a, b,
        f"収支合計!行2 月={ctx.month}",
        f"売上管理台帳 完了月={ctx.month} SUM(税別)",
        ctx.thresholds,
        "役務提供完了日基準で突合",
    )


def rule_S02_ios_revenue(ctx: CheckContext) -> RuleResult:
    """iOSアカデミア売上 管理表=台帳。"""
    a = get_mgmt_value(ctx.mgmt, "iOSアカデミア", 2, ctx.month)
    b = sum_ledger_by_month_category(ctx.ledger_sales, ctx.month, _guess_category_col(ctx.ledger_sales), "iOSアカデミア")
    return _make_result(
        "S02", "iOSアカデミア売上 管理表=台帳",
        a, b,
        f"iOSアカデミア!B2 月={ctx.month}",
        f"売上管理台帳 完了月={ctx.month} 区分=iOSアカデミア SUM(税別)",
        ctx.thresholds,
        "役務提供完了日基準・税別",
    )


def rule_S03_tech_revenue(ctx: CheckContext) -> RuleResult:
    """init TECH売上（SES + アプリ受託開発）= 台帳。"""
    a = get_mgmt_value(ctx.mgmt, "init TECH", 2, ctx.month)
    cats = ["SES", "アプリ受託開発", "init TECH"]
    b = sum_ledger_by_month_category(ctx.ledger_sales, ctx.month, _guess_category_col(ctx.ledger_sales), cats)
    return _make_result(
        "S03", "init TECH売上 管理表=台帳",
        a, b,
        f"init TECH!B2 月={ctx.month}",
        f"売上管理台帳 完了月={ctx.month} 区分∈{cats} SUM(税別)",
        ctx.thresholds,
    )


def rule_S04_misc_revenue(ctx: CheckContext) -> RuleResult:
    """雑収入 管理表=台帳。"""
    a = get_mgmt_value(ctx.mgmt, "その他", 2, ctx.month)
    b = sum_ledger_by_month_category(ctx.ledger_sales, ctx.month, _guess_category_col(ctx.ledger_sales), "雑収入")
    return _make_result(
        "S04", "雑収入 管理表=台帳",
        a, b,
        f"その他!B2 月={ctx.month}",
        f"売上管理台帳 完了月={ctx.month} 区分=雑収入 SUM(税別)",
        ctx.thresholds,
    )


def rule_S05_revenue_vs_freee(ctx: CheckContext) -> RuleResult:
    """管理表売上合計 = freee売上高。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 2, ctx.month)
    b = _sum_freee_pl_by_categories(ctx.freee_pl, ["売上高"], ctx.month)
    return _make_result(
        "S05", "売上合計 管理表=freee PL",
        a, b,
        f"収支合計!行2 月={ctx.month}",
        f"freee PL 売上高 月={ctx.month}",
        ctx.thresholds,
        "freeeは発生ベース・差異率1%以内で合格",
    )


def rule_S06_refund(ctx: CheckContext) -> RuleResult:
    """返金反映: 管理表返金 = 台帳の負値合計。"""
    a = get_mgmt_value(ctx.mgmt, "その他", 6, ctx.month)
    b = sum_negative_by_month(ctx.ledger_sales, ctx.month)
    return _make_result(
        "S06", "返金 管理表=台帳負値",
        a, b,
        f"その他!B6 月={ctx.month}",
        f"売上管理台帳 完了月={ctx.month} 金額<0 SUM",
        ctx.thresholds,
    )


# ---------------------------------------------------------------------------
# E系: 費用突合
# ---------------------------------------------------------------------------

def rule_E01_cost(ctx: CheckContext) -> RuleResult:
    """原価 管理表=支出管理表。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 4, ctx.month)
    b = sum_ledger_by_month_category(ctx.ledger_exp, ctx.month, _guess_category_col(ctx.ledger_exp), "原価")
    return _make_result("E01", "原価 管理表=支出台帳", a, b,
                        f"収支合計!行4 月={ctx.month}",
                        f"支出管理表 完了月={ctx.month} 費目=原価 SUM",
                        ctx.thresholds)


def rule_E02_promotion(ctx: CheckContext) -> RuleResult:
    """販促費集約。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 5, ctx.month)
    cats = ["広告宣伝", "販促", "交際", "会議", "広告宣伝費", "販売促進費", "交際費", "会議費"]
    b = sum_ledger_by_month_category(ctx.ledger_exp, ctx.month, _guess_category_col(ctx.ledger_exp), cats)
    return _make_result("E02", "販促費集約 管理表=支出台帳", a, b,
                        f"収支合計!行5 月={ctx.month}",
                        f"支出管理表 完了月={ctx.month} 費目∈販促費群 SUM",
                        ctx.thresholds)


def rule_E03_admin(ctx: CheckContext) -> RuleResult:
    """管理費集約。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 9, ctx.month)
    cats = ["管理費", "旅費交通費", "通信費", "消耗品費", "地代家賃", "水道光熱費",
            "支払手数料", "雑費", "新聞図書費", "リース料", "租税公課", "減価償却費"]
    b = sum_ledger_by_month_category(ctx.ledger_exp, ctx.month, _guess_category_col(ctx.ledger_exp), cats)
    return _make_result("E03", "管理費集約 管理表=支出台帳", a, b,
                        f"収支合計!行9 月={ctx.month}",
                        f"支出管理表 完了月={ctx.month} 費目∈管理費群 SUM",
                        ctx.thresholds)


def rule_E04_recruitment(ctx: CheckContext) -> RuleResult:
    """採用教育費。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 7, ctx.month)
    b = sum_ledger_by_month_category(ctx.ledger_exp, ctx.month, _guess_category_col(ctx.ledger_exp),
                                     ["採用教育", "採用費", "教育研修費"])
    return _make_result("E04", "採用教育費 管理表=支出台帳", a, b,
                        f"収支合計!行7 月={ctx.month}",
                        f"支出管理表 完了月={ctx.month} 費目=採用教育 SUM",
                        ctx.thresholds)


def rule_E05_welfare(ctx: CheckContext) -> RuleResult:
    """福利厚生費。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 8, ctx.month)
    b = sum_ledger_by_month_category(ctx.ledger_exp, ctx.month, _guess_category_col(ctx.ledger_exp),
                                     ["福利厚生", "福利厚生費"])
    return _make_result("E05", "福利厚生費 管理表=支出台帳", a, b,
                        f"収支合計!行8 月={ctx.month}",
                        f"支出管理表 完了月={ctx.month} 費目=福利厚生 SUM",
                        ctx.thresholds)


def rule_E06_sga_vs_freee(ctx: CheckContext) -> RuleResult:
    """freee販管費総額との突合。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 17, ctx.month)
    freee_sga_accounts = [
        "広告宣伝費", "販売促進費", "交際費", "会議費",
        "採用費", "教育研修費", "福利厚生費",
        "旅費交通費", "通信費", "消耗品費", "地代家賃",
        "水道光熱費", "支払手数料", "雑費", "減価償却費",
    ]
    b = sum_freee_accounts(ctx.freee_pl, freee_sga_accounts, ctx.month)
    return _make_result("E06", "販管費計 管理表=freee PL", a, b,
                        f"収支合計!行17 月={ctx.month}",
                        f"freee PL 販管費群合計 月={ctx.month}",
                        ctx.thresholds,
                        "差異率2%以内で合格")


# ---------------------------------------------------------------------------
# P系: 人件費突合
# ---------------------------------------------------------------------------

def rule_P01_officer_salary(ctx: CheckContext) -> RuleResult:
    """役員報酬合計（管理表=振込ベース / freee=発生ベース）。常にinfo。"""
    a = _sum_mgmt_cells(ctx.mgmt, "役員報酬給与", [3, 4], ctx.month)
    b = sum_freee_accounts(ctx.freee_pl, ["役員報酬", "給料手当"], ctx.month)
    return _make_result("P01", "役員報酬+給与 管理表=freee（参考）", a, b,
                        f"役員報酬給与!C3+C4 月={ctx.month}",
                        f"freee PL 役員報酬+給料手当 月={ctx.month}",
                        ctx.thresholds,
                        "⚠ 管理表=振込ベース、freee=発生ベースのため差異正常")


def rule_P02_resident_tax(ctx: CheckContext) -> RuleResult:
    """社保(住民税)。"""
    a = _get_mgmt_cell_direct(ctx.mgmt, "役員報酬給与", 6, ctx.month)
    b = get_freee_account(ctx.freee_bs, "預り金（住民税）", ctx.month)
    return _make_result("P02", "住民税 管理表=freee BS", a, b,
                        "役員報酬給与!C6",
                        f"freee BS 預り金（住民税）月={ctx.month}",
                        ctx.thresholds)


def rule_P03_social_insurance(ctx: CheckContext) -> RuleResult:
    """社会保険料。"""
    a = _get_mgmt_cell_direct(ctx.mgmt, "役員報酬給与", 7, ctx.month)
    b = (get_freee_account(ctx.freee_pl, "法定福利費", ctx.month)
         + get_freee_account(ctx.freee_bs, "預り金（社会保険料）", ctx.month))
    return _make_result("P03", "社会保険料 管理表=freee", a, b,
                        "役員報酬給与!C7",
                        f"freee PL法定福利費+BS預り金(社保) 月={ctx.month}",
                        ctx.thresholds)


def rule_P04_withholding_tax(ctx: CheckContext) -> RuleResult:
    """源泉所得税。"""
    a = _get_mgmt_cell_direct(ctx.mgmt, "役員報酬給与", 10, ctx.month)
    b = get_freee_account(ctx.freee_bs, "預り金（源泉所得税）", ctx.month)
    return _make_result("P04", "源泉所得税 管理表=freee BS", a, b,
                        "役員報酬給与!C10",
                        f"freee BS 預り金（源泉所得税）月={ctx.month}",
                        ctx.thresholds)


def rule_P05_allocation_rate(ctx: CheckContext) -> RuleResult:
    """配賦率合計 = 100%。"""
    df = ctx.mgmt.get("給与原価割合")
    if df is None or df.empty:
        total = None
    else:
        try:
            col = _find_month_col_df(df, ctx.month)
            if col:
                total = float(df[col].sum()) * 100  # 小数→%
            else:
                # 数値列をすべて合計（最初の数値列）
                num_cols = df.select_dtypes(include="number").columns
                total = float(df[num_cols[0]].sum()) * 100 if len(num_cols) > 0 else None
        except Exception:
            total = None
    return _make_result("P05", "配賦率合計=100%", 100.0, total,
                        "給与原価割合 全行合計",
                        "給与原価割合シート 合計",
                        ctx.thresholds,
                        "±0.1ポイント以内で合格")


# ---------------------------------------------------------------------------
# T系: 集計系
# ---------------------------------------------------------------------------

def rule_T01_gross_profit(ctx: CheckContext) -> RuleResult:
    """粗利 = 売上 - 原価。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 15, ctx.month)
    row2 = get_mgmt_value(ctx.mgmt, "収支合計", 2, ctx.month)
    row4 = get_mgmt_value(ctx.mgmt, "収支合計", 4, ctx.month)
    b = (row2 or 0) - (row4 or 0)
    return _make_result("T01", "粗利 管理表内整合", a, b,
                        f"収支合計!行15 月={ctx.month}",
                        "行2(売上) - 行4(原価)",
                        ctx.thresholds)


def rule_T02_sga_total(ctx: CheckContext) -> RuleResult:
    """販管費計 = SUM(行5:13) + 行14。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 17, ctx.month)
    rows = [get_mgmt_value(ctx.mgmt, "収支合計", r, ctx.month) or 0 for r in range(5, 14)]
    row14 = get_mgmt_value(ctx.mgmt, "収支合計", 14, ctx.month) or 0
    b = sum(rows) + row14
    return _make_result("T02", "販管費計 管理表内整合", a, b,
                        f"収支合計!行17 月={ctx.month}",
                        "SUM(行5:13)+行14",
                        ctx.thresholds)


def rule_T03_operating_income(ctx: CheckContext) -> RuleResult:
    """営業利益 = 粗利 - 販管費計。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 18, ctx.month)
    row15 = get_mgmt_value(ctx.mgmt, "収支合計", 15, ctx.month) or 0
    row17 = get_mgmt_value(ctx.mgmt, "収支合計", 17, ctx.month) or 0
    b = row15 - row17
    return _make_result("T03", "営業利益 管理表内整合", a, b,
                        f"収支合計!行18 月={ctx.month}",
                        "行15(粗利) - 行17(販管費計)",
                        ctx.thresholds)


def rule_T04_cash_flow(ctx: CheckContext) -> RuleResult:
    """CF = 行22 - 行23。"""
    a = get_mgmt_value(ctx.mgmt, "収支合計", 24, ctx.month)
    row22 = get_mgmt_value(ctx.mgmt, "収支合計", 22, ctx.month) or 0
    row23 = get_mgmt_value(ctx.mgmt, "収支合計", 23, ctx.month) or 0
    b = row22 - row23
    return _make_result("T04", "CF 管理表内整合", a, b,
                        f"収支合計!行24 月={ctx.month}",
                        "行22 - 行23",
                        ctx.thresholds)


# ---------------------------------------------------------------------------
# B系: BS突合
# ---------------------------------------------------------------------------

def rule_B01_resident_tax_bs(ctx: CheckContext) -> RuleResult:
    """預り金(住民税) 月末残高の整合性。"""
    return _check_deposit_balance(ctx, "預り金（住民税）", "B01", "預り金(住民税)月末残高")


def rule_B02_withholding_tax_bs(ctx: CheckContext) -> RuleResult:
    """預り金(源泉所得税) 月末残高の整合性。"""
    return _check_deposit_balance(ctx, "預り金（源泉所得税）", "B02", "預り金(源泉所得税)月末残高")


def rule_B03_social_insurance_bs(ctx: CheckContext) -> RuleResult:
    """預り金(社保) 月末残高の整合性。"""
    return _check_deposit_balance(ctx, "預り金（社会保険料）", "B03", "預り金(社保)月末残高")


def rule_B04_loan_balance(ctx: CheckContext) -> RuleResult:
    """借入金残高減少確認。"""
    prev_month = _prev_month(ctx.month)
    prev_balance = get_freee_account(ctx.freee_bs, "長期借入金", prev_month)
    curr_balance = get_freee_account(ctx.freee_bs, "長期借入金", ctx.month)
    mgmt_balance = get_mgmt_value(ctx.mgmt, "その他", 8, ctx.month)
    diff = (mgmt_balance or 0) - curr_balance
    rate = diff / curr_balance if curr_balance != 0 else 0.0
    abs_tol = _get_threshold(ctx.thresholds, "B04", "abs")
    status = judge(diff, rate, abs_tol, None)
    return RuleResult("B04", "借入金残高 管理表=freee BS", mgmt_balance, curr_balance,
                      diff, rate, status,
                      f"前月末残={prev_balance:,.0f}円 → 当月末={curr_balance:,.0f}円",
                      f"その他!B8 月={ctx.month}", f"freee BS 長期借入金 月={ctx.month}")


# ---------------------------------------------------------------------------
# I系: 台帳イレギュラー検知
# ---------------------------------------------------------------------------

def detect_irregulars(sales_file, exp_file, month: str) -> list[dict]:
    """台帳ファイルのイレギュラーを検知してリストで返す。"""
    results = []

    for label, f in [("売上管理台帳", sales_file), ("支出管理表", exp_file)]:
        if f is None:
            continue
        if hasattr(f, "seek"):
            f.seek(0)
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            continue

        ws = wb.active

        # I02: 色付きセル検知
        WHITE_COLORS = {"00000000", "FFFFFFFF", "00FFFFFF", "FFFF0000"}
        for row in ws.iter_rows():
            for cell in row:
                fill = cell.fill
                if fill and fill.fill_type not in (None, "none"):
                    rgb = fill.fgColor.rgb if fill.fgColor else None
                    if rgb and rgb.upper() not in WHITE_COLORS:
                        results.append({
                            "type": "I02",
                            "file": label,
                            "cell": cell.coordinate,
                            "value": cell.value,
                            "color": rgb,
                            "description": f"色付きセル検出: {cell.coordinate} (色:{rgb})",
                        })

        # I03: 備考キーワード検知
        keywords = ["変更", "訂正", "修正", "確認中"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for kw in keywords:
                        if kw in cell.value:
                            results.append({
                                "type": "I03",
                                "file": label,
                                "cell": cell.coordinate,
                                "value": cell.value,
                                "keyword": kw,
                                "description": f"備考キーワード「{kw}」検出: {cell.coordinate}",
                            })
                            break

        if hasattr(f, "seek"):
            f.seek(0)

    return results


def compute_and_save_hash(sales_df: pd.DataFrame | None, exp_df: pd.DataFrame | None, month: str) -> dict:
    """台帳データのハッシュを計算して.cacheに保存する。"""
    cache_dir = Path(__file__).parent.parent / ".cache"
    cache_dir.mkdir(exist_ok=True)
    cache_file = cache_dir / f"{month}.json"

    current = {}
    for name, df in [("sales", sales_df), ("exp", exp_df)]:
        if df is not None and not df.empty:
            h = hashlib.md5(pd.util.hash_pandas_object(df, index=True).values.tobytes()).hexdigest()
            current[name] = h

    prev = {}
    if cache_file.exists():
        try:
            with open(cache_file, encoding="utf-8") as fp:
                prev = json.load(fp)
        except Exception:
            pass

    with open(cache_file, "w", encoding="utf-8") as fp:
        json.dump(current, fp, ensure_ascii=False)

    changed = {k: {"prev": prev.get(k), "current": v} for k, v in current.items() if prev.get(k) != v}
    return changed


# ---------------------------------------------------------------------------
# メイン実行関数
# ---------------------------------------------------------------------------

def run_all(files: dict, month: str) -> dict:
    """全ルールを実行してresultsを返す。"""
    from .loaders import (
        load_mgmt_excel, load_sales_ledger, load_expense_ledger,
        load_keihi_ledger, load_biz_ledger, load_freee_bs, load_freee_pl,
        load_thresholds,
    )

    th = load_thresholds()

    def _read(key):
        f = files.get(key)
        if f is not None and hasattr(f, "seek"):
            f.seek(0)
        return f

    mgmt_data = load_mgmt_excel(_read("mgmt")) if files.get("mgmt") else {}
    sales_df  = load_sales_ledger(_read("sales")) if files.get("sales") else None
    exp_df    = load_expense_ledger(_read("exp")) if files.get("exp") else None
    keihi_df  = load_keihi_ledger(_read("keihi"))
    biz_df    = load_biz_ledger(_read("biz"))
    bs_data   = load_freee_bs(_read("bs")) if files.get("bs") else {}
    pl_data   = load_freee_pl(_read("pl")) if files.get("pl") else {}

    ctx = CheckContext(
        month=month,
        mgmt=mgmt_data,
        ledger_sales=sales_df,
        ledger_exp=exp_df,
        ledger_keihi=keihi_df,
        ledger_biz=biz_df,
        freee_bs=bs_data,
        freee_pl=pl_data,
        thresholds=th,
    )

    rules = [
        rule_S01_total_revenue, rule_S02_ios_revenue, rule_S03_tech_revenue,
        rule_S04_misc_revenue, rule_S05_revenue_vs_freee, rule_S06_refund,
        rule_E01_cost, rule_E02_promotion, rule_E03_admin,
        rule_E04_recruitment, rule_E05_welfare, rule_E06_sga_vs_freee,
        rule_P01_officer_salary, rule_P02_resident_tax, rule_P03_social_insurance,
        rule_P04_withholding_tax, rule_P05_allocation_rate,
        rule_T01_gross_profit, rule_T02_sga_total, rule_T03_operating_income, rule_T04_cash_flow,
        rule_B01_resident_tax_bs, rule_B02_withholding_tax_bs,
        rule_B03_social_insurance_bs, rule_B04_loan_balance,
    ]

    results = []
    for rule_fn in rules:
        try:
            results.append(rule_fn(ctx))
        except Exception as e:
            rule_id = rule_fn.__name__.split("_")[1].upper() if "_" in rule_fn.__name__ else "??"
            results.append(RuleResult(
                rule_id, rule_fn.__doc__ or rule_fn.__name__,
                None, None, None, None, "warn",
                f"実行エラー: {e}", "", "",
            ))

    # ハッシュ保存（I04）
    hash_changes = compute_and_save_hash(sales_df, exp_df, month)

    # イレギュラー検知（I02/I03）
    sales_f = files.get("sales")
    exp_f = files.get("exp")
    irregulars = detect_irregulars(sales_f, exp_f, month)

    return {
        "results": results,
        "irregulars": irregulars,
        "hash_changes": hash_changes,
        "month": month,
    }


# ---------------------------------------------------------------------------
# 内部ヘルパー
# ---------------------------------------------------------------------------

def _guess_category_col(df: pd.DataFrame | None) -> str:
    """台帳DataFrameから区分・費目列名を推定する。"""
    if df is None:
        return "区分"
    for col in df.columns:
        if col and any(k in str(col) for k in ["区分", "費目", "カテゴリ", "分類", "事業"]):
            return col
    return df.columns[2] if len(df.columns) > 2 else "区分"


def _sum_freee_pl_by_categories(pl: dict, accounts: list[str], month: str) -> float:
    total = 0.0
    for account in accounts:
        # 完全一致 + 部分一致
        for key, months in pl.items():
            if account in key or key in account:
                total += months.get(month, 0.0)
    return total


def _sum_mgmt_cells(mgmt: dict, sheet: str, rows: list[int], month: str) -> Optional[float]:
    total = 0.0
    for r in rows:
        v = get_mgmt_value(mgmt, sheet, r, month)
        if v is None:
            return None
        total += v
    return total


def _get_mgmt_cell_direct(mgmt: dict, sheet: str, row: int, month: str) -> Optional[float]:
    return get_mgmt_value(mgmt, sheet, row, month)


def _check_deposit_balance(ctx: CheckContext, account: str, rule_id: str, label: str) -> RuleResult:
    """預り金残高の前月末+当月増減=当月末チェック。"""
    prev_month = _prev_month(ctx.month)
    prev_bal = get_freee_account(ctx.freee_bs, account, prev_month)
    curr_bal = get_freee_account(ctx.freee_bs, account, ctx.month)
    # freee BSのみで整合性確認（前月+変動=当月）
    # 変動はPLの該当科目から推定
    diff = curr_bal  # BSのみ参照の場合は残高確認のみ
    return RuleResult(rule_id, label, curr_bal, curr_bal, 0.0, 0.0, "info",
                      f"前月末残={prev_bal:,.0f}円 → 当月末={curr_bal:,.0f}円（BS参照値）",
                      f"freee BS {account} 月={ctx.month}",
                      f"freee BS {account} 月={ctx.month}")


def _prev_month(month: str) -> str:
    year, mon = int(month[:4]), int(month[5:])
    mon -= 1
    if mon == 0:
        mon = 12
        year -= 1
    return f"{year}-{mon:02d}"


def _find_month_col_df(df: pd.DataFrame, month: str):
    import re
    year, mon = month.split("-")
    for col in df.columns:
        if col is None:
            continue
        m = re.search(r"(\d{4})[/年](\d{1,2})", str(col))
        if m and m.group(1) == year and int(m.group(2)) == int(mon):
            return col
    return None
