"""各ファイルの読み込み関数群。"""
import io
import re
from pathlib import Path
from typing import Optional, Union

import pandas as pd
import openpyxl
import pdfplumber
import yaml


# ---------------------------------------------------------------------------
# 収益管理表
# ---------------------------------------------------------------------------

def load_mgmt_excel(file_like) -> dict:
    """
    収益管理表を読み込む。
    戻り値: {sheet_name: DataFrame or dict}
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    result = {}

    target_sheets = [
        "収支合計",
        "iOSアカデミア",
        "init TECH",
        "その他",
        "役員報酬給与",
        "給与原価割合",
        "売上管理 請求書",
        "売上管理 計算式",
        "支出管理 請求書",
        "支出管理 計算式",
    ]

    for sheet_name in wb.sheetnames:
        for target in target_sheets:
            if target in sheet_name:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    result[target] = pd.DataFrame(rows[1:], columns=rows[0])
                else:
                    result[target] = pd.DataFrame()
                break

    return result


def list_months(file_like) -> list[str]:
    """収益管理表のヘッダ行から月リストを取得する。"""
    if file_like is None:
        return []
    try:
        # Streamlit UploadedFile は read() 後に巻き戻す
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        wb = openpyxl.load_workbook(file_like, data_only=True, read_only=True)
        ws = wb["収支合計"] if "収支合計" in wb.sheetnames else wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        months = []
        for val in header:
            if val is None:
                continue
            s = str(val)
            # "2026/01" "2026年1月" などを YYYY-MM に正規化
            m = re.search(r"(\d{4})[/年](\d{1,2})", s)
            if m:
                months.append(f"{m.group(1)}-{int(m.group(2)):02d}")
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        return sorted(set(months))
    except Exception:
        return []


def get_mgmt_value(mgmt: dict, sheet: str, row: int, month: str) -> Optional[float]:
    """
    管理表の特定シート・行・月の値を取得する。
    month: "YYYY-MM"
    row: 1始まりの行番号（ヘッダ除く）
    """
    df = mgmt.get(sheet)
    if df is None or df.empty:
        return None
    try:
        col = _find_month_col(df, month)
        if col is None:
            return None
        val = df.iloc[row - 1][col]
        return float(val) if val is not None else None
    except (IndexError, ValueError, TypeError):
        return None


def _find_month_col(df: pd.DataFrame, month: str) -> Optional[str]:
    """DataFrameのカラムから対象月に対応する列名を探す。"""
    year, mon = month.split("-")
    for col in df.columns:
        if col is None:
            continue
        s = str(col)
        m = re.search(r"(\d{4})[/年](\d{1,2})", s)
        if m and m.group(1) == year and int(m.group(2)) == int(mon):
            return col
    return None


# ---------------------------------------------------------------------------
# 売上管理台帳
# ---------------------------------------------------------------------------

def load_sales_ledger(file_like) -> pd.DataFrame:
    """
    売上管理台帳を読み込む。
    H列: 役務提供完了日 → 完了月(YYYY-MM)列を追加
    """
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    df = pd.read_excel(file_like, header=0)

    # H列（index 7）を完了日として解釈
    date_col = df.columns[7] if len(df.columns) > 7 else None
    if date_col is not None:
        df["完了月"] = pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%Y-%m")
        df.rename(columns={date_col: "役務提供完了日"}, inplace=True)

    # 税別金額列を標準化（E列 = index 4）
    amount_col = df.columns[4] if len(df.columns) > 4 else None
    if amount_col is not None:
        df.rename(columns={amount_col: "税別"}, inplace=True)
        df["税別"] = pd.to_numeric(df["税別"], errors="coerce").fillna(0)

    return df


# ---------------------------------------------------------------------------
# 支出管理表
# ---------------------------------------------------------------------------

def load_expense_ledger(file_like) -> pd.DataFrame:
    """
    支出管理表を読み込む。
    K列: 役務提供完了日 → 完了月(YYYY-MM)列を追加
    """
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    df = pd.read_excel(file_like, header=0)

    date_col = df.columns[10] if len(df.columns) > 10 else None
    if date_col is not None:
        df["完了月"] = pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%Y-%m")
        df.rename(columns={date_col: "役務提供完了日"}, inplace=True)

    amount_col = df.columns[4] if len(df.columns) > 4 else None
    if amount_col is not None:
        df.rename(columns={amount_col: "税別"}, inplace=True)
        df["税別"] = pd.to_numeric(df["税別"], errors="coerce").fillna(0)

    return df


# ---------------------------------------------------------------------------
# 経費台帳
# ---------------------------------------------------------------------------

def load_keihi_ledger(file_like) -> pd.DataFrame:
    if file_like is None:
        return pd.DataFrame()
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    df = pd.read_excel(file_like, header=0)
    # 計上日列を自動検出
    for col in df.columns:
        if "計上日" in str(col) or "日付" in str(col):
            df["完了月"] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m")
            break
    return df


# ---------------------------------------------------------------------------
# 事業別支出管理
# ---------------------------------------------------------------------------

def load_biz_ledger(file_like) -> pd.DataFrame:
    if file_like is None:
        return pd.DataFrame()
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    df = pd.read_excel(file_like, header=0)
    return df


# ---------------------------------------------------------------------------
# freee BS（試算表）
# ---------------------------------------------------------------------------

def load_freee_bs(file_like) -> dict:
    """
    freee 貸借対照表(試算表)を読み込む。
    戻り値: {科目名: {月: 金額}}
    """
    if file_like is None:
        return {}
    if hasattr(file_like, "seek"):
        file_like.seek(0)

    name = getattr(file_like, "name", "")
    if name.lower().endswith(".pdf"):
        return _parse_freee_bs_pdf(file_like)
    else:
        return _parse_freee_bs_excel(file_like)


def _parse_freee_bs_excel(file_like) -> dict:
    df = pd.read_excel(file_like, header=0, index_col=0)
    result = {}
    for account, row in df.iterrows():
        if pd.isna(account):
            continue
        result[str(account)] = {}
        for col in df.columns:
            m = re.search(r"(\d{4})[/年](\d{1,2})", str(col))
            if m:
                month_key = f"{m.group(1)}-{int(m.group(2)):02d}"
                val = row[col]
                result[str(account)][month_key] = float(val) if pd.notna(val) else 0.0
    return result


def _parse_freee_bs_pdf(file_like) -> dict:
    """PDFからBSテーブルを抽出する（pdfplumber使用）。"""
    result = {}
    data = file_like.read() if hasattr(file_like, "read") else open(file_like, "rb").read()
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                header = table[0]
                for row in table[1:]:
                    if not row or not row[0]:
                        continue
                    account = str(row[0]).strip()
                    result[account] = {}
                    for i, col_name in enumerate(header[1:], 1):
                        if col_name is None or i >= len(row):
                            continue
                        m = re.search(r"(\d{4})[/年](\d{1,2})", str(col_name))
                        if m:
                            month_key = f"{m.group(1)}-{int(m.group(2)):02d}"
                            try:
                                val = float(str(row[i]).replace(",", "").replace("△", "-").replace("▲", "-"))
                            except (ValueError, TypeError):
                                val = 0.0
                            result[account][month_key] = val
    return result


# ---------------------------------------------------------------------------
# freee PL（月次推移表）
# ---------------------------------------------------------------------------

def load_freee_pl(file_like) -> dict:
    """
    freee 損益計算書(月次推移表)を読み込む。
    戻り値: {科目名: {月: 金額}}
    """
    if file_like is None:
        return {}
    if hasattr(file_like, "seek"):
        file_like.seek(0)

    name = getattr(file_like, "name", "")
    if name.lower().endswith(".pdf"):
        return _parse_freee_pl_pdf(file_like)
    else:
        return _parse_freee_pl_excel(file_like)


def _parse_freee_pl_excel(file_like) -> dict:
    df = pd.read_excel(file_like, header=0, index_col=0)
    result = {}
    for account, row in df.iterrows():
        if pd.isna(account):
            continue
        result[str(account)] = {}
        for col in df.columns:
            m = re.search(r"(\d{4})[/年](\d{1,2})", str(col))
            if m:
                month_key = f"{m.group(1)}-{int(m.group(2)):02d}"
                val = row[col]
                result[str(account)][month_key] = float(val) if pd.notna(val) else 0.0
    return result


def _parse_freee_pl_pdf(file_like) -> dict:
    result = {}
    data = file_like.read() if hasattr(file_like, "read") else open(file_like, "rb").read()
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                header = table[0]
                for row in table[1:]:
                    if not row or not row[0]:
                        continue
                    account = str(row[0]).strip()
                    result[account] = {}
                    for i, col_name in enumerate(header[1:], 1):
                        if col_name is None or i >= len(row):
                            continue
                        m = re.search(r"(\d{4})[/年](\d{1,2})", str(col_name))
                        if m:
                            month_key = f"{m.group(1)}-{int(m.group(2)):02d}"
                            try:
                                val = float(str(row[i]).replace(",", "").replace("△", "-").replace("▲", "-"))
                            except (ValueError, TypeError):
                                val = 0.0
                            result[account][month_key] = val
    return result


# ---------------------------------------------------------------------------
# 設定ファイル
# ---------------------------------------------------------------------------

def load_thresholds(path: Optional[str] = None) -> dict:
    if path is None:
        path = Path(__file__).parent.parent / "config" / "thresholds.yaml"
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_account_map(path: Optional[str] = None) -> dict:
    if path is None:
        path = Path(__file__).parent.parent / "config" / "account_map.yaml"
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


if __name__ == "__main__":
    # 動作確認用
    th = load_thresholds()
    print("thresholds loaded:", list(th.keys()))
    am = load_account_map()
    print("account_map loaded:", list(am.keys())[:5])
